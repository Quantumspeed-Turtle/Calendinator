import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import tkinter as tk
from tkinter import simpledialog
import os


class DateInputDialog(tk.simpledialog.Dialog):
    def __init__(self, parent, title):
        self.start_date_str = None
        self.end_date_str = None
        super().__init__(parent, title)

    def body(self, frame):
        tk.Label(frame, text="Start Date (YYYY-MM-DD):").grid(row=0, column=0, padx=5, pady=5)
        tk.Label(frame, text="End Date (YYYY-MM-DD):").grid(row=1, column=0, padx=5, pady=5)

        self.start_date_entry = tk.Entry(frame)
        self.start_date_entry.grid(row=0, column=1, padx=5, pady=5)

        self.end_date_entry = tk.Entry(frame)
        self.end_date_entry.grid(row=1, column=1, padx=5, pady=5)

    def apply(self):
        self.start_date_str = self.start_date_entry.get()
        self.end_date_str = self.end_date_entry.get()


class CalendarGenerator:
    def __init__(self, start_date, end_date):
        self.start_date = start_date
        self.end_date = end_date
        self.calendar_data = self.generate_calendar(start_date, end_date)
        

    # Function to generate a calendar for the specified date range
    def generate_calendar(self, start_date, end_date):
        df_list = []

        current_date = start_date

        if current_date.weekday() != 0:
            current_date += timedelta(days=(7 - current_date.weekday()))

        while current_date <= end_date:
            week_number = current_date.strftime('%W')
            week_days = [current_date.strftime('%d.%m')]

            for _ in range(4):
                current_date += timedelta(days=1)
                week_days.append(current_date.strftime('%d.%m'))

            df_list.append(pd.DataFrame({'CW': week_number, 'MONDAY': week_days[0], 'TUESDAY': week_days[1],
                                        'WEDNESDAY': week_days[2], 'THURSDAY': week_days[3], 'FRIDAY': week_days[4]}, index=[0]))

            current_date += timedelta(days=3)

        df = pd.concat(df_list, ignore_index=True)
        return df


    def create_excel_file(self, file_name):

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            self.calendar_data.to_excel(writer, index=False, sheet_name='Calendar')

            for i in range(2, 7):
                column_letter = openpyxl.utils.get_column_letter(i)
                writer.sheets['Calendar'].column_dimensions[column_letter].width = 28

                date_format = 'DD/MM'
                for cell in writer.sheets['Calendar'][column_letter]:
                    if cell.row > 1:
                        cell.number_format = date_format

                header_cell = writer.sheets['Calendar'].cell(row=1, column=i)
                header_cell.font = Font(bold=True, name='Arial')

            for row in writer.sheets['Calendar'].iter_rows():
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                    cell.font = openpyxl.styles.Font(size=10, name='Arial')
                    writer.sheets['Calendar'].row_dimensions[cell.row].height = 25
                    if cell.row == 1:
                        cell.font = openpyxl.styles.Font(size=14, bold=True, name='Arial')



    def create_pdf_file(self, pdf_file_name):

        plt.figure(figsize=(8.27, 11.69))
        plt.axis('off')
        column_widths = [0.08, 0.2, 0.2, 0.2, 0.2, 0.2]
        table = plt.table(cellText=self.calendar_data.values, colLabels=self.calendar_data.columns, cellLoc='left', loc='center',
                          colWidths=column_widths)

         # Add the title with the years to the PDF
        #if start_year == end_year:
        #    title = f"{start_year}"
        #    plt.text(1.2, 1.257, title, fontsize=18, fontweight='bold', ha='left', va='center')

        #else:
        #    title = f"{start_year}/{end_year}"
        #    plt.text(1.1, 1.257, title, fontsize=18, fontweight='bold', ha='left', va='center')


        table.auto_set_font_size(False)
        table.set_fontsize(7)

        for col in range(len(self.calendar_data.columns)):
            cell = table[0, col]
            cell._text.set_fontweight('bold')
            cell.set_fontsize(13)

        
        for row in range(1, len(self.calendar_data)+1):
            cell = table[row, 0]
            cell._text.set_fontweight('bold')

        table.scale(1.5, 1.5)
        plt.subplots_adjust(top=1)

        plt.savefig(pdf_file_name, bbox_inches='tight', pad_inches=0.9)
        plt.close()

def get_user_input():
    root = tk.Tk()

    # Center the GUI window on the screen
    # Center the GUI window on the screen
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    # Define the amount to move the window to the left and up
    move_left = 200
    move_up = 50
    # Calculate the new X and Y coordinates for the window
    x = screen_width // 2 - window_width // 2 - move_left
    y = screen_height // 2 - window_height // 2 - move_up

    root.geometry(f"+{x}+{y}")
    root.withdraw()

    dialog = DateInputDialog(root, "Enter Start and End Dates")

    if dialog.start_date_str and dialog.end_date_str:
        try:
            start_date = datetime.strptime(dialog.start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(dialog.end_date_str, "%Y-%m-%d")
        except ValueError:
            print("Invalid date format. Please use the YYYY-MM-DD format.")
            return None, None

        return start_date, end_date
    else:
        return None, None



if __name__ == "__main__":
    start_date, end_date = get_user_input()

    if start_date is not None and end_date is not None:
        calendar_generator = CalendarGenerator(start_date, end_date)

        if start_date == end_date:
            file_name = 'calendar_' + start_date.strftime("%Y-%m-%d")+'.xlsx'
            pdf_file_name = 'calendar_' + start_date.strftime("%Y-%m-%d")+'.pdf'
        else:
            file_name = 'calendar_'+start_date.strftime("%Y-%m-%d")+'_'+end_date.strftime("%Y-%m-%d")+'.xlsx'
            pdf_file_name = 'calendar_'+start_date.strftime("%Y-%m-%d")+'_'+end_date.strftime("%Y-%m-%d")+'.pdf'


        # Create the Excel file
        calendar_generator.create_excel_file(file_name)
        print(f"Calendar created and saved as '{file_name}'.")


        # Create the PDF file
        calendar_generator.create_pdf_file(pdf_file_name)
        print(f"PDF created and saved as '{pdf_file_name}'.")
